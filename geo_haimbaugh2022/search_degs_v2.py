import openpyxl
import re
import csv
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Build Ensembl to gene name mapping
mapping = {}
with open(os.path.join(SCRIPT_DIR, 'ensembl_to_gene.tsv'), 'r') as f:
    next(f)  # skip header
    for line in f:
        parts = line.strip().split('\t')
        if len(parts) == 2 and parts[1]:
            mapping[parts[0]] = parts[1]

# Define target gene patterns
targets = {
    'HOX genes': re.compile(r'^hox[abcd]', re.IGNORECASE),
    'phf8': re.compile(r'^phf8[ab]?$', re.IGNORECASE),
    'kdm6b': re.compile(r'^kdm6b[ab]?$', re.IGNORECASE),
    'kdm2b': re.compile(r'^kdm2b[ab]?$', re.IGNORECASE),
    'kdm7a/7b': re.compile(r'^kdm7[ab][ab]?$', re.IGNORECASE),
    'ezh': re.compile(r'^ezh[12]$', re.IGNORECASE),
    'dnmt': re.compile(r'^dnmt', re.IGNORECASE),
    'sox': re.compile(r'^sox(5|9|9[ab]|10|11[ab]?)', re.IGNORECASE),
    'pax': re.compile(r'^pax[37][ab]?$', re.IGNORECASE),
    'twist': re.compile(r'^twist[12][ab]?$', re.IGNORECASE),
    'foxd3': re.compile(r'^foxd3$', re.IGNORECASE),
    'tbx': re.compile(r'^tbx[23][ab]?$', re.IGNORECASE),
}

target_ensembl = {}
target_category = {}
for eid, gname in mapping.items():
    for category, pattern in targets.items():
        if pattern.match(gname):
            target_ensembl[eid] = gname
            target_category[eid] = category
            break

# Search all three DEG files
results = []
files = [
    ('Table S3_F0 DEGs.xlsx', 'F0'),
    ('Table S4_F1 DEGs.xlsx', 'F1'),
    ('Table S5_F2 DEGs.xlsx', 'F2'),
]

for fname, generation in files:
    wb = openpyxl.load_workbook(os.path.join(SCRIPT_DIR, fname), read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            eid = str(row[0]) if row[0] else ''
            if eid in target_ensembl:
                try:
                    log2fc = float(row[2]) if row[2] is not None else None
                except (ValueError, TypeError):
                    log2fc = None
                try:
                    pvalue = float(row[5]) if row[5] is not None else None
                except (ValueError, TypeError):
                    pvalue = None
                try:
                    padj = float(row[6]) if row[6] is not None else None
                except (ValueError, TypeError):
                    padj = None
                results.append({
                    'category': target_category[eid],
                    'generation': generation,
                    'compound': sheet_name,
                    'gene': target_ensembl[eid],
                    'ensembl_id': eid,
                    'log2FC': log2fc,
                    'pvalue': pvalue,
                    'padj': padj,
                })
    wb.close()

# Save all results to CSV
results.sort(key=lambda x: (x['category'], x['gene'], x['generation'], x['compound']))
with open(os.path.join(SCRIPT_DIR, 'target_gene_hits_all.csv'), 'w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=['category', 'gene', 'ensembl_id', 'generation', 'compound', 'log2FC', 'pvalue', 'padj'])
    writer.writeheader()
    writer.writerows(results)

# Save significant results to CSV
sig_results = [r for r in results if r['padj'] is not None and r['padj'] < 0.05]
with open(os.path.join(SCRIPT_DIR, 'target_gene_hits_significant.csv'), 'w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=['category', 'gene', 'ensembl_id', 'generation', 'compound', 'log2FC', 'pvalue', 'padj'])
    writer.writeheader()
    writer.writerows(sig_results)

# Save nominally significant results to CSV
nom_sig = [r for r in results if r['pvalue'] is not None and r['pvalue'] < 0.05]
with open(os.path.join(SCRIPT_DIR, 'target_gene_hits_nominal_p05.csv'), 'w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=['category', 'gene', 'ensembl_id', 'generation', 'compound', 'log2FC', 'pvalue', 'padj'])
    writer.writeheader()
    writer.writerows(nom_sig)

print(f"Total hits across all conditions: {len(results)}")
print(f"Significant (padj < 0.05): {len(sig_results)}")
print(f"Nominally significant (p < 0.05): {len(nom_sig)}")
print("CSV files saved.")
